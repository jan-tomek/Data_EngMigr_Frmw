import openpyxl 
import argparse
import pandas
import csv
import subprocess
import os
import tempfile
import sys

sys.stdout.reconfigure(encoding='utf-8')

import codecs
codecs.register_error("strict", codecs.ignore_errors)

parser = argparse.ArgumentParser(description='Loads Excel to database')
parser.add_argument("--infile", required=True, help="Source Excel to be loaded.")
parser.add_argument("--lookup", action="store_true", help="Load lookup individually located in each Excel.")
parser.add_argument("--manual", action="store_true", help="Load manual data Excels.")
parser.add_argument("--codetest", action="store_true", help="Generate ddl for code test.")

print("#######################")
print("# ExcelLoader started #")
print("#######################")

args = parser.parse_args()

inFilePath = args.infile
lookup = args.lookup
manual = args.manual
codetest = args.codetest

if lookup: print("Lookup mode.")
if manual: print("Manual data Excel mode.")
if codetest: print("Generate ddl for code test.")

print("Input file path: " + inFilePath)

workbook = openpyxl.load_workbook(filename = inFilePath, data_only=True)
for worksheet in workbook.worksheets:
    if worksheet.sheet_state == 'hidden': continue

    if lookup: table = os.path.basename(inFilePath).replace('#','/').replace('.xlsx','')
    else: table = worksheet.title
    print("Table name: ["+table+']')

    sqlFilePath = inFilePath.replace('.xlsx','')+'#'+worksheet.title+'.sql'
    sqlFile = open(sqlFilePath, 'w', encoding='utf-8') 
    print("SQL file path: " + sqlFilePath)
    
    if lookup: dbase = 'OUT_LKP'
    else: dbase = 'IN_MAN'

    headerRow = 0

    # Find first data row 
    for row in worksheet.iter_rows():
        if row[0].value != None: break
        headerRow = headerRow + 1
    print("First data row: "+str(headerRow + 1))

    ddl  = 'DROP TABLE IF EXISTS ['+dbase+'].['+table+'];'
    ddl += '\nGO'
    ddl += '\nCREATE TABLE ['+dbase+'].['+table+'] ('
    
    # Read header
    numCols = 1
    colsList=[]
    number = 1
    for col in worksheet.iter_cols(max_row=32):
        print('.',end = '')
        sys.stdout.flush()
        column = str(col[headerRow].value)
        if not column: break
        if column == 'None':             # No header
             column = column + str(number)
             number = number + 1
        if column in colsList:           # Dup
             column = column + str(number)
             number = number + 1
        colsList.append(column)
        if col[headerRow].col_idx == 1: ddl += ('\n     ')
        else: ddl += ('\n    ,')
        ddl += '['+column+'] NVARCHAR(4000) NOT NULL'
        numCols = numCols + 1

    ddl += '\n);'
    ddl += '\nGO'
    if codetest: 
        print(ddl.replace('['+dbase+']',"[@ENV@#CodeTest]."+'['+dbase+']')+'\n',file = sqlFile)    
    print(ddl,file = sqlFile)

    print(' - '+str(numCols) + ' columns')
    
    if lookup: 
        print('\nDROP VIEW IF EXISTS [OUT].['+table+'];',file = sqlFile)
        print('GO',file = sqlFile)
        print('CREATE VIEW [OUT].['+table+'] AS',file = sqlFile)
        print('SELECT * FROM ['+dbase+'].['+table+'];',file = sqlFile)
        print('GO',file = sqlFile)
    print('',file = sqlFile)
   
    rowNum = 0
    for row in worksheet.iter_rows(min_row=headerRow+2):
        if row[0].value == None: continue

        print('INSERT INTO ['+dbase+'].['+table+'] VALUES(',file = sqlFile,end = '')

        for cell in row:
            if cell.col_idx > 1: print(',',file = sqlFile,end = '')
            cval = str(cell.value).replace("'","''").replace('=TRUE()','1').replace('=FALSE()','0') 

            print("N'"+cval.replace('None','')+"'",file = sqlFile,end = '')

        rowNum = cell.row - headerRow - 1
        print('); -- '+str(rowNum),file = sqlFile)

    print("Rows exported: " + str(rowNum))
   
    sqlFile.close() 

workbook.close

